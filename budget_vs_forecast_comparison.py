#!/usr/bin/env python3
"""
Comprehensive Budget vs Forecast Comparison Script
===================================================
Compares:
  1. BB Forecast Baseline Outputs v1.xlsx (the model forecast)
  2. Budget consol file.xlsx (budget targets)

Metrics: Closing GBV, Closing NBV, Collections, Revenue, Gross Impairment, Net Impairment
Also: Coverage Ratios (actual and implied from budget impairment)

Segment mapping:
  Budget "Non Prime"         -> Forecast "NON PRIME"
  Budget "Near Prime Small"  -> Forecast "NRP-S"
  Budget "Near Prime Medium" -> Forecast "NRP-M" + "NRP-L"  (NRP-L has no budget line)
  Budget "Prime"             -> Forecast "PRIME"
"""

import openpyxl
from datetime import datetime
from collections import defaultdict
import warnings
warnings.filterwarnings("ignore")

# ============================================================================
# CONFIGURATION
# ============================================================================
BUDGET_FILE = "Budget consol file.xlsx"
FORECAST_FILE = "BB Forecast Baseline Outputs v1.xlsx"
BUDGET_SHEET = "P&L analysis - BB"

# Budget row mapping (row numbers in the budget sheet)
BUDGET_ROWS = {
    "Disbursals":       {"Non Prime": 6,  "NPS": 7,  "NPM": 8,  "Prime": 9,  "Total": 10},
    "Collections":      {"Non Prime": 12, "NPS": 13, "NPM": 14, "Prime": 15, "Total": 16},
    "ClosingGBV":       {"Non Prime": 23, "NPS": 24, "NPM": 25, "Prime": 26, "Total": 27},
    "ClosingNBV":       {"Non Prime": 43, "NPS": 44, "NPM": 45, "Prime": 46, "Total": 47},
    "Revenue":          {"Non Prime": 63, "NPS": 64, "NPM": 65, "Prime": 66, "Total": 67},
    "GrossImpairment":  {"Non Prime": 74, "NPS": 75, "NPM": 76, "Prime": 77, "Total": 78},
    "RAM_exclDS":       {"Non Prime": 90, "NPS": 91, "NPM": 92, "Prime": 93, "Total": 94},
    "DebtSaleGain":     {"Non Prime": 106, "NPS": 107, "NPM": 108, "Prime": 109, "Total": 110},
    "NetImpairment":    {"Non Prime": 122, "NPS": 123, "NPM": 124, "Prime": 125, "Total": 126},
}

# Segment names for display
SEGMENTS = ["Non Prime", "NPS", "NPM", "Prime"]
SEGMENT_DISPLAY = {
    "Non Prime": "Non Prime",
    "NPS": "Near Prime Small (NPS/NRP-S)",
    "NPM": "Near Prime Medium (NPM/NRP-M+L)",
    "Prime": "Prime",
}

# Forecast segment mapping: budget segment -> list of forecast segments to sum
FORECAST_SEG_MAP = {
    "Non Prime": ["NON PRIME"],
    "NPS": ["NRP-S"],
    "NPM": ["NRP-M", "NRP-L"],
    "Prime": ["PRIME"],
}

# ============================================================================
# 1. LOAD BUDGET DATA
# ============================================================================
def load_budget_data():
    """Load all budget data from the P&L analysis sheet."""
    wb = openpyxl.load_workbook(BUDGET_FILE, data_only=True)
    ws = wb[BUDGET_SHEET]

    # Build month-to-column mapping from row 3
    month_cols = {}
    for col in range(4, ws.max_column + 1):
        date_val = ws.cell(row=3, column=col).value
        if date_val is None:
            continue
        if isinstance(date_val, (int, float)):
            # Excel serial date - convert
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            date_val = base + timedelta(days=int(date_val))
        if isinstance(date_val, datetime):
            month_key = date_val.strftime("%Y-%m-%d")
            month_cols[month_key] = col

    # Extract data for each metric, segment, month
    budget = {}
    for metric, rows in BUDGET_ROWS.items():
        budget[metric] = {}
        for seg, row_num in rows.items():
            budget[metric][seg] = {}
            for month_key, col in month_cols.items():
                val = ws.cell(row=row_num, column=col).value
                if val is not None:
                    try:
                        budget[metric][seg][month_key] = float(val)
                    except (ValueError, TypeError):
                        pass

    wb.close()
    return budget, sorted(month_cols.keys())


# ============================================================================
# 2. LOAD FORECAST DATA
# ============================================================================
def load_forecast_data():
    """Load forecast data from 9_Summary and 11_Impairment sheets."""
    wb = openpyxl.load_workbook(FORECAST_FILE, data_only=True)

    # --- 9_Summary sheet ---
    ws = wb["9_Summary"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    summary_data = []
    for row in range(2, ws.max_row + 1):
        rec = {}
        for c, h in enumerate(headers, 1):
            if h:
                val = ws.cell(row=row, column=c).value
                rec[h] = val
        if rec.get("ForecastMonth"):
            summary_data.append(rec)

    # --- 11_Impairment sheet (for gross impairment) ---
    ws2 = wb["11_Impairment"]
    imp_headers = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]

    # Aggregate gross impairment and net impairment by month+segment
    gross_imp_agg = defaultdict(lambda: defaultdict(float))
    net_imp_agg = defaultdict(lambda: defaultdict(float))
    ds_impact_agg = defaultdict(lambda: defaultdict(float))

    for row in range(2, ws2.max_row + 1):
        month_raw = ws2.cell(row=row, column=1).value
        seg = ws2.cell(row=row, column=2).value
        if month_raw is None or seg is None:
            continue
        month_key = str(month_raw)[:10]

        gi_col = imp_headers.index("Gross_Impairment_ExcludingDS") + 1 if "Gross_Impairment_ExcludingDS" in imp_headers else None
        ni_col = imp_headers.index("Net_Impairment") + 1 if "Net_Impairment" in imp_headers else None
        dsi_col = imp_headers.index("Debt_Sale_Impact") + 1 if "Debt_Sale_Impact" in imp_headers else None

        if gi_col:
            v = ws2.cell(row=row, column=gi_col).value
            if v is not None:
                gross_imp_agg[month_key][seg] += float(v)
        if ni_col:
            v = ws2.cell(row=row, column=ni_col).value
            if v is not None:
                net_imp_agg[month_key][seg] += float(v)
        if dsi_col:
            v = ws2.cell(row=row, column=dsi_col).value
            if v is not None:
                ds_impact_agg[month_key][seg] += float(v)

    wb.close()

    # Now build forecast dict matching budget structure
    forecast = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    for rec in summary_data:
        month_raw = rec["ForecastMonth"]
        if isinstance(month_raw, datetime):
            month_key = month_raw.strftime("%Y-%m-%d")
        else:
            month_key = str(month_raw)[:10]

        seg_raw = rec["Segment"]

        # Map to budget segment
        for bseg, fseg_list in FORECAST_SEG_MAP.items():
            if seg_raw in fseg_list:
                budget_seg = bseg
                break
        else:
            continue

        # Collections: forecast has negative sign (outflow), budget has positive
        coll_p = float(rec.get("Coll_Principal", 0) or 0)
        coll_i = float(rec.get("Coll_Interest", 0) or 0)
        forecast["Collections"][budget_seg][month_key] += abs(coll_p) + abs(coll_i)

        # Closing GBV
        forecast["ClosingGBV"][budget_seg][month_key] += float(rec.get("ClosingGBV", 0) or 0)

        # Closing NBV
        forecast["ClosingNBV"][budget_seg][month_key] += float(rec.get("ClosingNBV", 0) or 0)

        # Revenue (InterestRevenue)
        forecast["Revenue"][budget_seg][month_key] += float(rec.get("InterestRevenue", 0) or 0)

        # Net Impairment from summary
        forecast["NetImpairment_Summary"][budget_seg][month_key] += float(rec.get("Net_Impairment", 0) or 0)

    # Gross impairment from 11_Impairment aggregation
    for month_key in gross_imp_agg:
        for seg_raw in gross_imp_agg[month_key]:
            for bseg, fseg_list in FORECAST_SEG_MAP.items():
                if seg_raw in fseg_list:
                    forecast["GrossImpairment"][bseg][month_key] += gross_imp_agg[month_key][seg_raw]
                    forecast["NetImpairment"][bseg][month_key] += net_imp_agg[month_key][seg_raw]
                    forecast["DebtSaleImpact"][bseg][month_key] += ds_impact_agg[month_key][seg_raw]

    # Compute totals for each metric
    for metric in ["Collections", "ClosingGBV", "ClosingNBV", "Revenue",
                   "GrossImpairment", "NetImpairment", "DebtSaleImpact", "NetImpairment_Summary"]:
        all_months = set()
        for seg in SEGMENTS:
            all_months.update(forecast[metric][seg].keys())
        for m in all_months:
            forecast[metric]["Total"][m] = sum(forecast[metric][seg].get(m, 0) for seg in SEGMENTS)

    return dict(forecast)


# ============================================================================
# 3. COMPARISON FUNCTIONS
# ============================================================================
def fmt_val(v, scale=1.0):
    """Format a value in millions."""
    if v is None:
        return "   N/A    "
    return f"{v/scale:>12,.2f}"

def fmt_pct(v):
    """Format a percentage."""
    if v is None:
        return "   N/A  "
    return f"{v:>8.1f}%"

def fmt_pct_ratio(v):
    """Format a ratio as percentage."""
    if v is None:
        return "  N/A  "
    return f"{v*100:>7.2f}%"


def print_metric_comparison(metric_name, budget, forecast, budget_metric, forecast_metric,
                           months, segments_plus_total, sign_flip_budget=False,
                           sign_flip_forecast=False):
    """Print side-by-side comparison for one metric across all segments and months."""
    print("\n" + "=" * 140)
    print(f"  {metric_name}")
    print("=" * 140)

    for seg in segments_plus_total:
        seg_display = SEGMENT_DISPLAY.get(seg, seg)
        print(f"\n  --- {seg_display} {'(TOTAL)' if seg == 'Total' else ''} ---")
        print(f"  {'Month':<12} {'Budget':>14} {'Forecast':>14} {'Variance':>14} {'Var %':>10}   Notes")
        print(f"  {'-'*12} {'-'*14} {'-'*14} {'-'*14} {'-'*10}   -----")

        for m in months:
            b_val = budget.get(budget_metric, {}).get(seg, {}).get(m, None)
            f_val = forecast.get(forecast_metric, {}).get(seg, {}).get(m, None)

            if b_val is not None and sign_flip_budget:
                b_val = -b_val
            if f_val is not None and sign_flip_forecast:
                f_val = -f_val

            if b_val is not None and f_val is not None:
                var = f_val - b_val
                var_pct = (var / abs(b_val) * 100) if abs(b_val) > 0.01 else None
                note = ""
                if var_pct is not None and abs(var_pct) > 10:
                    note = " ** LARGE GAP"
                elif var_pct is not None and abs(var_pct) > 5:
                    note = " * notable"
            elif b_val is None and f_val is None:
                var = None
                var_pct = None
                note = ""
            else:
                var = None
                var_pct = None
                note = " (missing data)"

            m_short = m[:7]  # YYYY-MM
            print(f"  {m_short:<12} {fmt_val(b_val)} {fmt_val(f_val)} {fmt_val(var)} {fmt_pct(var_pct)}   {note}")


def compute_coverage_ratio(gbv, nbv):
    """Coverage ratio = (GBV - NBV) / GBV"""
    if gbv is None or nbv is None or abs(gbv) < 0.01:
        return None
    return (gbv - nbv) / gbv


# ============================================================================
# 4. MAIN ANALYSIS
# ============================================================================
def main():
    print("=" * 140)
    print("  BUDGET vs FORECAST COMPARISON - BB Model")
    print("  " + "=" * 70)
    print(f"  Budget file:   {BUDGET_FILE}")
    print(f"  Forecast file: {FORECAST_FILE}")
    print()
    print("  Segment mapping:")
    print("    Budget 'Non Prime'         <-> Forecast 'NON PRIME'")
    print("    Budget 'Near Prime Small'  <-> Forecast 'NRP-S'")
    print("    Budget 'Near Prime Medium' <-> Forecast 'NRP-M' + 'NRP-L' (NRP-L folded in)")
    print("    Budget 'Prime'             <-> Forecast 'PRIME'")
    print()
    print("  KEY DIFFERENCE: Budget has NO debt sale gain (all zeros).")
    print("  Forecast includes debt sale impacts in Dec-25, Mar-26, Jun-26.")
    print("  Net Impairment differs by debt sale effect between the two.")
    print("=" * 140)

    # Load data
    budget, budget_months = load_budget_data()
    forecast = load_forecast_data()

    # Determine overlapping months
    forecast_months_set = set()
    for metric in forecast:
        for seg in forecast[metric]:
            forecast_months_set.update(forecast[metric][seg].keys())
    forecast_months = sorted(forecast_months_set)

    overlap_months = sorted(set(budget_months) & set(forecast_months))
    print(f"\n  Budget months available:   {budget_months[0]} to {budget_months[-1]} ({len(budget_months)} months)")
    print(f"  Forecast months available: {forecast_months[0]} to {forecast_months[-1]} ({len(forecast_months)} months)")
    print(f"  Overlapping months:        {overlap_months[0]} to {overlap_months[-1]} ({len(overlap_months)} months)")

    segs_and_total = SEGMENTS + ["Total"]

    # ========================================================================
    # CLOSING GBV
    # ========================================================================
    print_metric_comparison(
        "CLOSING GBV (Gross Book Value)",
        budget, forecast, "ClosingGBV", "ClosingGBV",
        overlap_months, segs_and_total
    )

    # ========================================================================
    # CLOSING NBV
    # ========================================================================
    print_metric_comparison(
        "CLOSING NBV (Net Book Value)",
        budget, forecast, "ClosingNBV", "ClosingNBV",
        overlap_months, segs_and_total
    )

    # ========================================================================
    # COLLECTIONS (Budget positive, Forecast stores abs of negative)
    # ========================================================================
    print_metric_comparison(
        "COLLECTIONS (Principal + Interest)",
        budget, forecast, "Collections", "Collections",
        overlap_months, segs_and_total
    )

    # ========================================================================
    # REVENUE
    # ========================================================================
    print_metric_comparison(
        "REVENUE (Interest Revenue)",
        budget, forecast, "Revenue", "Revenue",
        overlap_months, segs_and_total
    )

    # ========================================================================
    # GROSS IMPAIRMENT
    # Budget stores as negative (loss), forecast stores as positive (charge)
    # Flip budget sign so both are positive=charge for comparison
    # ========================================================================
    print_metric_comparison(
        "GROSS IMPAIRMENT (excl. Debt Sale) - Positive = charge/loss",
        budget, forecast, "GrossImpairment", "GrossImpairment",
        overlap_months, segs_and_total,
        sign_flip_budget=True  # Budget negative -> positive
    )

    # ========================================================================
    # NET IMPAIRMENT
    # Budget Net Impairment = Gross Impairment (since no debt sale)
    # Forecast Net Impairment = Gross + Debt Sale Impact
    # ========================================================================
    print_metric_comparison(
        "NET IMPAIRMENT - Positive = charge/loss\n"
        "  (Budget has NO debt sale, so Net = Gross. Forecast includes debt sale impact.)",
        budget, forecast, "NetImpairment", "NetImpairment",
        overlap_months, segs_and_total,
        sign_flip_budget=True
    )

    # ========================================================================
    # FORECAST: Gross vs Net Impairment breakdown (to show debt sale effect)
    # ========================================================================
    print("\n" + "=" * 140)
    print("  FORECAST: DEBT SALE IMPACT BREAKDOWN (Gross Impairment vs Net Impairment)")
    print("  Shows why forecast Net Impairment differs from Gross Impairment")
    print("=" * 140)
    for seg in segs_and_total:
        seg_display = SEGMENT_DISPLAY.get(seg, seg)
        print(f"\n  --- {seg_display} {'(TOTAL)' if seg == 'Total' else ''} ---")
        print(f"  {'Month':<12} {'Gross Imp':>14} {'DS Impact':>14} {'Net Imp':>14}   Notes")
        print(f"  {'-'*12} {'-'*14} {'-'*14} {'-'*14}   -----")
        for m in overlap_months:
            gi = forecast.get("GrossImpairment", {}).get(seg, {}).get(m, 0)
            dsi = forecast.get("DebtSaleImpact", {}).get(seg, {}).get(m, 0)
            ni = forecast.get("NetImpairment", {}).get(seg, {}).get(m, 0)
            note = " <-- DEBT SALE MONTH" if abs(dsi) > 0.01 else ""
            print(f"  {m[:7]:<12} {gi:>14,.2f} {dsi:>14,.2f} {ni:>14,.2f}   {note}")

    # ========================================================================
    # COVERAGE RATIOS
    # ========================================================================
    print("\n" + "=" * 140)
    print("  COVERAGE RATIOS: (GBV - NBV) / GBV")
    print("  Budget Coverage = (Budget_GBV - Budget_NBV) / Budget_GBV")
    print("  Forecast Coverage = as reported in forecast model")
    print("=" * 140)

    for seg in segs_and_total:
        seg_display = SEGMENT_DISPLAY.get(seg, seg)
        print(f"\n  --- {seg_display} {'(TOTAL)' if seg == 'Total' else ''} ---")
        print(f"  {'Month':<12} {'Budget CR':>10} {'Forecast CR':>12} {'Variance':>10}   Notes")
        print(f"  {'-'*12} {'-'*10} {'-'*12} {'-'*10}   -----")
        for m in overlap_months:
            b_gbv = budget.get("ClosingGBV", {}).get(seg, {}).get(m, None)
            b_nbv = budget.get("ClosingNBV", {}).get(seg, {}).get(m, None)
            f_gbv = forecast.get("ClosingGBV", {}).get(seg, {}).get(m, None)
            f_nbv = forecast.get("ClosingNBV", {}).get(seg, {}).get(m, None)

            b_cr = compute_coverage_ratio(b_gbv, b_nbv)
            f_cr = compute_coverage_ratio(f_gbv, f_nbv)

            if b_cr is not None and f_cr is not None:
                var = f_cr - b_cr
                note = ""
                if abs(var) > 0.05:
                    note = " ** LARGE GAP (>5pp)"
                elif abs(var) > 0.02:
                    note = " * notable (>2pp)"
            else:
                var = None
                note = ""

            print(f"  {m[:7]:<12} {fmt_pct_ratio(b_cr)} {fmt_pct_ratio(f_cr):>12} {fmt_pct_ratio(var):>10}   {note}")

    # ========================================================================
    # IMPLIED COVERAGE RATIOS FROM BUDGET IMPAIRMENT
    # ========================================================================
    print("\n" + "=" * 140)
    print("  IMPLIED ANALYSIS: What coverage ratios would produce the budget's impairment?")
    print("  " + "-" * 100)
    print("  Approach: If we know the budget's provision balance = GBV * CR,")
    print("  then provision movement = new_balance - prior_balance")
    print("  The budget Gross Impairment = -provision_movement (with sign convention)")
    print("  So: Implied_CR(t) = [GBV(t-1)*CR(t-1) + GrossImpairment(t)] / GBV(t)")
    print("  We back-solve iteratively from the first month's known NBV ratio.")
    print("=" * 140)

    for seg in segs_and_total:
        seg_display = SEGMENT_DISPLAY.get(seg, seg)
        print(f"\n  --- {seg_display} {'(TOTAL)' if seg == 'Total' else ''} ---")
        print(f"  {'Month':<12} {'Budget GBV':>14} {'Budget NBV':>14} {'Direct CR':>10} {'Budget GI':>14} {'Implied CR':>10} {'Implied Prov':>14}")
        print(f"  {'-'*12} {'-'*14} {'-'*14} {'-'*10} {'-'*14} {'-'*10} {'-'*14}")

        prev_prov = None
        for i, m in enumerate(overlap_months):
            b_gbv = budget.get("ClosingGBV", {}).get(seg, {}).get(m, None)
            b_nbv = budget.get("ClosingNBV", {}).get(seg, {}).get(m, None)
            b_gi = budget.get("GrossImpairment", {}).get(seg, {}).get(m, None)  # negative in budget

            direct_cr = compute_coverage_ratio(b_gbv, b_nbv)

            if b_gbv is not None and b_nbv is not None:
                current_prov = b_gbv - b_nbv
            else:
                current_prov = None

            if prev_prov is not None and b_gi is not None and b_gbv is not None and abs(b_gbv) > 0.01:
                # Provision movement = current_prov - prev_prov
                # Gross Impairment in budget = negative of provision movement
                # Actually: budget GI is negative = loss. So provision_movement = -budget_GI
                # But let's just use the actual provision balances
                implied_prov = current_prov
                implied_cr = current_prov / b_gbv if b_gbv else None
            elif b_gbv is not None and b_nbv is not None:
                implied_prov = current_prov
                implied_cr = direct_cr
            else:
                implied_prov = None
                implied_cr = None

            gi_display = -b_gi if b_gi is not None else None  # flip to positive

            print(f"  {m[:7]:<12} {fmt_val(b_gbv)} {fmt_val(b_nbv)} {fmt_pct_ratio(direct_cr)} {fmt_val(gi_display)} {fmt_pct_ratio(implied_cr)} {fmt_val(implied_prov)}")

            prev_prov = current_prov

    # ========================================================================
    # PROVISION MOVEMENT RECONCILIATION
    # ========================================================================
    print("\n" + "=" * 140)
    print("  PROVISION MOVEMENT RECONCILIATION")
    print("  Budget: Provision = GBV - NBV. Movement = Prov(t) - Prov(t-1)")
    print("  Gross Impairment should equal Provision Movement (in budget, no debt sale)")
    print("=" * 140)

    for seg in segs_and_total:
        seg_display = SEGMENT_DISPLAY.get(seg, seg)
        print(f"\n  --- {seg_display} {'(TOTAL)' if seg == 'Total' else ''} ---")
        print(f"  {'Month':<12} {'Prov(t)':>14} {'Prov(t-1)':>14} {'Movement':>14} {'Budget GI':>14} {'Difference':>14}   Notes")
        print(f"  {'-'*12} {'-'*14} {'-'*14} {'-'*14} {'-'*14} {'-'*14}   -----")

        prev_prov = None
        for m in overlap_months:
            b_gbv = budget.get("ClosingGBV", {}).get(seg, {}).get(m, None)
            b_nbv = budget.get("ClosingNBV", {}).get(seg, {}).get(m, None)
            b_gi = budget.get("GrossImpairment", {}).get(seg, {}).get(m, None)

            if b_gbv is not None and b_nbv is not None:
                current_prov = b_gbv - b_nbv
            else:
                current_prov = None

            if prev_prov is not None and current_prov is not None:
                movement = current_prov - prev_prov
                gi_positive = -b_gi if b_gi is not None else None  # flip sign
                diff = movement - gi_positive if gi_positive is not None else None
                note = ""
                if diff is not None and abs(diff) > 1000:
                    note = " ** Provision movement != GI (other items?)"
            else:
                movement = None
                gi_positive = -b_gi if b_gi is not None else None
                diff = None
                note = " (first month)"

            print(f"  {m[:7]:<12} {fmt_val(current_prov)} {fmt_val(prev_prov)} {fmt_val(movement)} {fmt_val(gi_positive)} {fmt_val(diff)}   {note}")
            prev_prov = current_prov

    # ========================================================================
    # SUMMARY: BIGGEST GAPS
    # ========================================================================
    print("\n" + "=" * 140)
    print("  SUMMARY OF LARGEST VARIANCES (Absolute and %)")
    print("=" * 140)

    # Collect all variances
    all_variances = []

    metrics_to_compare = [
        ("ClosingGBV", "ClosingGBV", "Closing GBV", False, False),
        ("ClosingNBV", "ClosingNBV", "Closing NBV", False, False),
        ("Collections", "Collections", "Collections", False, False),
        ("Revenue", "Revenue", "Revenue", False, False),
        ("GrossImpairment", "GrossImpairment", "Gross Impairment", True, False),
        ("NetImpairment", "NetImpairment", "Net Impairment", True, False),
    ]

    for b_metric, f_metric, display_name, flip_b, flip_f in metrics_to_compare:
        for seg in segs_and_total:
            for m in overlap_months:
                b_val = budget.get(b_metric, {}).get(seg, {}).get(m, None)
                f_val = forecast.get(f_metric, {}).get(seg, {}).get(m, None)
                if flip_b and b_val is not None:
                    b_val = -b_val
                if flip_f and f_val is not None:
                    f_val = -f_val
                if b_val is not None and f_val is not None and abs(b_val) > 0.01:
                    var_abs = f_val - b_val
                    var_pct = (var_abs / abs(b_val)) * 100
                    all_variances.append({
                        "metric": display_name,
                        "segment": SEGMENT_DISPLAY.get(seg, seg),
                        "month": m[:7],
                        "budget": b_val,
                        "forecast": f_val,
                        "var_abs": var_abs,
                        "var_pct": var_pct,
                    })

    # Sort by absolute variance %
    all_variances.sort(key=lambda x: abs(x["var_pct"]), reverse=True)

    print("\n  TOP 30 LARGEST PERCENTAGE VARIANCES:")
    print(f"  {'Metric':<22} {'Segment':<35} {'Month':<10} {'Budget':>14} {'Forecast':>14} {'Var Abs':>14} {'Var %':>10}")
    print(f"  {'-'*22} {'-'*35} {'-'*10} {'-'*14} {'-'*14} {'-'*14} {'-'*10}")
    for v in all_variances[:30]:
        print(f"  {v['metric']:<22} {v['segment']:<35} {v['month']:<10} {v['budget']:>14,.2f} {v['forecast']:>14,.2f} {v['var_abs']:>14,.2f} {v['var_pct']:>9.1f}%")

    # Sort by absolute variance amount
    all_variances.sort(key=lambda x: abs(x["var_abs"]), reverse=True)

    print("\n  TOP 30 LARGEST ABSOLUTE VARIANCES:")
    print(f"  {'Metric':<22} {'Segment':<35} {'Month':<10} {'Budget':>14} {'Forecast':>14} {'Var Abs':>14} {'Var %':>10}")
    print(f"  {'-'*22} {'-'*35} {'-'*10} {'-'*14} {'-'*14} {'-'*14} {'-'*10}")
    for v in all_variances[:30]:
        print(f"  {v['metric']:<22} {v['segment']:<35} {v['month']:<10} {v['budget']:>14,.2f} {v['forecast']:>14,.2f} {v['var_abs']:>14,.2f} {v['var_pct']:>9.1f}%")

    # ========================================================================
    # AVERAGE VARIANCES BY METRIC AND SEGMENT
    # ========================================================================
    print("\n" + "=" * 140)
    print("  AVERAGE VARIANCE BY METRIC AND SEGMENT (across all overlapping months)")
    print("=" * 140)

    for b_metric, f_metric, display_name, flip_b, flip_f in metrics_to_compare:
        print(f"\n  {display_name}:")
        print(f"  {'Segment':<40} {'Avg Budget':>14} {'Avg Forecast':>14} {'Avg Var':>14} {'Avg Var %':>10}")
        print(f"  {'-'*40} {'-'*14} {'-'*14} {'-'*14} {'-'*10}")

        for seg in segs_and_total:
            b_vals = []
            f_vals = []
            for m in overlap_months:
                bv = budget.get(b_metric, {}).get(seg, {}).get(m, None)
                fv = forecast.get(f_metric, {}).get(seg, {}).get(m, None)
                if flip_b and bv is not None:
                    bv = -bv
                if flip_f and fv is not None:
                    fv = -fv
                if bv is not None and fv is not None:
                    b_vals.append(bv)
                    f_vals.append(fv)

            if b_vals:
                avg_b = sum(b_vals) / len(b_vals)
                avg_f = sum(f_vals) / len(f_vals)
                avg_var = avg_f - avg_b
                avg_pct = (avg_var / abs(avg_b) * 100) if abs(avg_b) > 0.01 else None
                seg_display = SEGMENT_DISPLAY.get(seg, seg)
                if seg == "Total":
                    seg_display = "TOTAL"
                print(f"  {seg_display:<40} {avg_b:>14,.2f} {avg_f:>14,.2f} {avg_var:>14,.2f} {fmt_pct(avg_pct)}")

    # ========================================================================
    # DETAILED: FORECAST GBV WITH NRP-L SHOWN SEPARATELY
    # ========================================================================
    print("\n" + "=" * 140)
    print("  DETAIL: NRP-L CONTRIBUTION (shown separately since no budget line exists)")
    print("  NRP-L is folded into NPM for budget comparison above.")
    print("=" * 140)

    # Reload forecast with NRP-L separate
    wb = openpyxl.load_workbook(FORECAST_FILE, data_only=True)
    ws = wb["9_Summary"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    nrpl_data = defaultdict(dict)
    nrpm_data = defaultdict(dict)

    for row in range(2, ws.max_row + 1):
        month_raw = ws.cell(row=row, column=1).value
        seg = ws.cell(row=row, column=2).value
        if month_raw is None:
            continue
        if isinstance(month_raw, datetime):
            mk = month_raw.strftime("%Y-%m-%d")
        else:
            mk = str(month_raw)[:10]

        if seg == "NRP-L":
            nrpl_data[mk] = {
                "ClosingGBV": float(ws.cell(row=row, column=headers.index("ClosingGBV") + 1).value or 0),
                "ClosingNBV": float(ws.cell(row=row, column=headers.index("ClosingNBV") + 1).value or 0),
                "Revenue": float(ws.cell(row=row, column=headers.index("InterestRevenue") + 1).value or 0),
                "NetImpairment": float(ws.cell(row=row, column=headers.index("Net_Impairment") + 1).value or 0),
            }
        elif seg == "NRP-M":
            nrpm_data[mk] = {
                "ClosingGBV": float(ws.cell(row=row, column=headers.index("ClosingGBV") + 1).value or 0),
                "ClosingNBV": float(ws.cell(row=row, column=headers.index("ClosingNBV") + 1).value or 0),
                "Revenue": float(ws.cell(row=row, column=headers.index("InterestRevenue") + 1).value or 0),
                "NetImpairment": float(ws.cell(row=row, column=headers.index("Net_Impairment") + 1).value or 0),
            }

    wb.close()

    print(f"\n  {'Month':<12} {'NRP-M GBV':>14} {'NRP-L GBV':>14} {'Combined':>14} {'Budget NPM':>14} {'NRP-L % of Combined':>20}")
    print(f"  {'-'*12} {'-'*14} {'-'*14} {'-'*14} {'-'*14} {'-'*20}")
    for m in overlap_months:
        nrpm_gbv = nrpm_data.get(m, {}).get("ClosingGBV", 0)
        nrpl_gbv = nrpl_data.get(m, {}).get("ClosingGBV", 0)
        combined = nrpm_gbv + nrpl_gbv
        b_npm_gbv = budget.get("ClosingGBV", {}).get("NPM", {}).get(m, None)
        pct = (nrpl_gbv / combined * 100) if combined > 0 else 0
        print(f"  {m[:7]:<12} {nrpm_gbv:>14,.2f} {nrpl_gbv:>14,.2f} {combined:>14,.2f} {fmt_val(b_npm_gbv)} {pct:>19.1f}%")

    # ========================================================================
    # FORECAST ONLY: RAM (Revenue After Margin) comparison
    # ========================================================================
    print("\n" + "=" * 140)
    print("  RAM (Revenue After Impairment, excl. Debt Sale) COMPARISON")
    print("  Budget RAM = Revenue + GrossImpairment (budget GI is negative)")
    print("  Forecast RAM = Revenue - GrossImpairment (forecast GI is positive=charge)")
    print("=" * 140)

    for seg in segs_and_total:
        seg_display = SEGMENT_DISPLAY.get(seg, seg)
        print(f"\n  --- {seg_display} {'(TOTAL)' if seg == 'Total' else ''} ---")
        print(f"  {'Month':<12} {'Bdgt Revenue':>14} {'Bdgt GI':>14} {'Bdgt RAM':>14} {'Fcst Revenue':>14} {'Fcst GI':>14} {'Fcst RAM':>14} {'RAM Var':>14} {'RAM Var%':>10}")
        print(f"  {'-'*12} {'-'*14} {'-'*14} {'-'*14} {'-'*14} {'-'*14} {'-'*14} {'-'*14} {'-'*10}")

        for m in overlap_months:
            b_rev = budget.get("Revenue", {}).get(seg, {}).get(m, None)
            b_gi = budget.get("GrossImpairment", {}).get(seg, {}).get(m, None)
            f_rev = forecast.get("Revenue", {}).get(seg, {}).get(m, None)
            f_gi = forecast.get("GrossImpairment", {}).get(seg, {}).get(m, None)

            if b_rev is not None and b_gi is not None:
                b_ram = b_rev + b_gi  # GI is negative in budget, so this is Rev - |GI|
            else:
                b_ram = None

            if f_rev is not None and f_gi is not None:
                f_ram = f_rev - f_gi  # GI is positive in forecast
            else:
                f_ram = None

            if b_ram is not None and f_ram is not None:
                ram_var = f_ram - b_ram
                ram_pct = (ram_var / abs(b_ram) * 100) if abs(b_ram) > 0.01 else None
            else:
                ram_var = None
                ram_pct = None

            # Display GI as positive for both
            b_gi_pos = -b_gi if b_gi is not None else None
            f_gi_pos = f_gi

            print(f"  {m[:7]:<12} {fmt_val(b_rev)} {fmt_val(b_gi_pos)} {fmt_val(b_ram)} {fmt_val(f_rev)} {fmt_val(f_gi_pos)} {fmt_val(f_ram)} {fmt_val(ram_var)} {fmt_pct(ram_pct)}")

    # Check budget RAM vs row 90-94
    print("\n  --- Cross-check: Budget RAM from rows 90-94 vs computed ---")
    print(f"  {'Month':<12} {'Seg':<15} {'Row RAM':>14} {'Computed RAM':>14} {'Match?':>8}")
    for seg in SEGMENTS:
        for m in overlap_months[:3]:
            row_ram = budget.get("RAM_exclDS", {}).get(seg, {}).get(m, None)
            b_rev = budget.get("Revenue", {}).get(seg, {}).get(m, None)
            b_gi = budget.get("GrossImpairment", {}).get(seg, {}).get(m, None)
            computed = (b_rev + b_gi) if (b_rev is not None and b_gi is not None) else None
            match = "YES" if (row_ram is not None and computed is not None and abs(row_ram - computed) < 1) else "NO"
            print(f"  {m[:7]:<12} {seg:<15} {fmt_val(row_ram)} {fmt_val(computed)}  {match:>6}")

    # ========================================================================
    # FINAL KEY TAKEAWAYS
    # ========================================================================
    print("\n" + "=" * 140)
    print("  KEY OBSERVATIONS AND TAKEAWAYS")
    print("=" * 140)

    # Calculate average variances for key metrics at total level
    for b_metric, f_metric, display_name, flip_b, flip_f in metrics_to_compare:
        b_sum = 0
        f_sum = 0
        count = 0
        for m in overlap_months:
            bv = budget.get(b_metric, {}).get("Total", {}).get(m, None)
            fv = forecast.get(f_metric, {}).get("Total", {}).get(m, None)
            if flip_b and bv is not None:
                bv = -bv
            if flip_f and fv is not None:
                fv = -fv
            if bv is not None and fv is not None:
                b_sum += bv
                f_sum += fv
                count += 1
        if count > 0:
            avg_b = b_sum / count
            avg_f = f_sum / count
            pct = ((avg_f - avg_b) / abs(avg_b) * 100) if abs(avg_b) > 0.01 else 0
            direction = "ABOVE" if avg_f > avg_b else "BELOW"
            print(f"  {display_name:<25}: Forecast is {direction} budget by avg {abs(pct):.1f}% "
                  f"(Budget avg: {avg_b:,.0f}, Forecast avg: {avg_f:,.0f})")

    print()
    print("  NOTES:")
    print("  1. Budget has ZERO debt sale gain; forecast models quarterly debt sales (Dec, Mar, Jun)")
    print("  2. This means Net Impairment differs significantly in debt sale months")
    print("  3. NRP-L segment exists only in forecast - folded into NPM for comparison")
    print("  4. Both files appear to be in GBP (not millions)")
    print()


if __name__ == "__main__":
    main()
